"""
    目标地址：http://www.chinatax.gov.cn/chinatax/n810346/n810825/index.html
    1. 采集
    采集以下栏目：
        增值税、消费税的前两页数据
    每个栏目的：
        标题、发文日期、文号

    2. 保存
        将所有信息保存到`税务局.xlsx`文件，依据栏目名创建数据表，每个栏目的数据保存到对应的数据表
        将所有信息分别保存到`税务局-增值税.json` `税务局-消费税.json`文件
        将所有信息分别保存到`税务局-增值税.csv` `税务局-消费税.json`文件
        将所有信息分别保存到`suiwujuzengzhi`数据表与`suiwujuxiaofei`数据表
        
        数据库相关参数：
            ip：134.175.188.27
            端口：3306
            用户：windows
            密码：123456
            数据库：test
"""
import requests
from openpyxl import Workbook
from itertools import chain
import json
import csv
url = 'http://www.chinatax.gov.cn/sfc/query.ejf?title=fgk&method=json'
headers ={
    'Cookie':'JSESSIONID=90E50F65B052DAA4B978FAEE1D919849; yfx_c_g_u_id_10003701=_ck20011219024516231176792065863; yfx_f_l_v_t_10003701=f_t_1578826965619__r_t_1578826965619__v_t_1578826965619__r_c_0; _Jo0OQK=5F61AD4BB9B06F867A5641BCBBBC40BD46ED90A096840F3CF627C1F916E51ED11B05E6511ADB5FF304ACDD22A23D904004BDCF5FF08D5C342FD2EE39FB375B6E56634275DAD340EB4DDFFF13AA80B4DD4EFFFF13AA80B4DD4EF13908CA53B564A60B4EE6E78E8917019GJ1Z1Kw==',
    'Host':'www.chinatax.gov.cn',
    'Origin':'http://www.chinatax.gov.cn',
    'Referer':'http://www.chinatax.gov.cn/chinatax/n810346/n810825/index.html',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36',
    'X-Requested-With':'XMLHttpRequest',
}
def get_data(tax,page):
    data = {
        'C2': tax,
        'timeOption': '0',
        'page': str(page),
        'pageSize': '15',
        'keyPlace': '1',
        'qt': '*',
        'sort': 'dateDesc',
    }
    return data
items1 =[]
items2 =[]
for i in range(1,3):
    res = requests.post(url,data=get_data('消费税',i),headers=headers)
    items1.extend(res.json()['resultList'])
    res = requests.post(url, data=get_data('增值税', i), headers=headers)
    items2.extend(res.json()['resultList'])
############################ 保存到excel表中去 ####################################################
# def save_data(tax,items):
#     wb = Workbook()
#     sheet = wb.create_sheet(tax)
#     sheet.cell(row=1, column=1).value = '标题'
#     sheet.cell(row=1, column=2).value = '发文日期'
#     sheet.cell(row=1, column=3).value = '文号'
#     num=2
#     for item in items:
#         sheet.cell(row=num, column=1).value = dreTitle = item['dreTitle']
#         sheet.cell(row=num, column=2).value = docDate = item['docDate']
#         sheet.cell(row=num, column=3).value = DOCNO = item['myValues']['DOCNO']
#         print(item['dreTitle'],item['docDate'],item['myValues']['DOCNO'])
#         num +=1
#     wb.save(tax + '.xlsx')
# save_data('消费税',items1)
# save_data('增值税',items2)
###################################保存到json文件中去 #################################
# with open('税务.json','a',encoding='utf-8') as f:
#     for item in chain(items1,items2):
#         shuju ={'dreTitle':item['dreTitle'],'docDate':item['docDate'],'DOCNO':item['myValues']['DOCNO']}
#         f.write(json.dumps(shuju,ensure_ascii=False)+"\n")
# ############################# 保存到csc 文件中去######################################
# with open("税务.csv",'a',encoding='utf-8',newline='') as f2:
#     for item in chain(items1,items2):
#         spamwriter = csv.writer(f2, delimiter=',')
#         shuju =[item['dreTitle'],item['docDate'],item['myValues']['DOCNO']]
#         spamwriter.writerow(shuju)
# ############################# 保存到mysql数据库中去######################################
#  暂时没写。。。。。。。。。。。。。。。。。。。。。。
### 基本流程  1》导包，2》创建conn连接对象，3》创建游标对象，3》插上能见表，4》插入数据，（必须保存，否则没用），将ccnn关闭
#    sql = "insert into 表名(列名)  values (%s,%s,%s)"
#   关键操作，cursor.excute(sql,[item['dreTitle'],item['docDate'],item['myValues']['DOCNO']])
#   conn.commit()
#  最后断开连接
###  备注：不喜欢用mysql


